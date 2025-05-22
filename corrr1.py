import pandas as pd
import numpy as np
from scipy.stats import spearmanr
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import scipy.stats as stats
stats.spearmanr
file_path = "data/selection.csv"

selection = pd.read_csv(file_path)

selection = selection.drop(columns=['pathogen load'])

selection.columns=selection.columns.str.replace('hwsd_soil_clm_res_awt_soc','awt_soc')
selection.columns=selection.columns.str.replace('hwsd_soil_clm_res_pct_clay','pct_clay')
selection.columns=selection.columns.str.replace('hwsd_soil_clm_res_dom_mu','dom_mu')
selection.columns=selection.columns.str.replace('hand_500m_china_03_08', 'hand')

selection.columns = selection.columns.str.replace('_', ' ')

corr_matrix = selection.corr(method='spearman')

print(corr_matrix)
p_matrix = selection.corr(method=lambda x, y: spearmanr(x, y)[1])


formatted_matrix = corr_matrix.copy()
for i in range(len(corr_matrix)):
    for j in range(len(corr_matrix.columns)):

        value = f"{corr_matrix.iloc[i, j]:.2f}"
        
        if corr_matrix.iloc[i, j] < 0:
            value = f"{value}"
        
        if p_matrix.iloc[i, j] < 0.05:
            value += "*"
        
        formatted_matrix.iloc[i, j] = value

mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=0)

lower_triangle_corr = formatted_matrix.where(~mask)

heatmap_data = lower_triangle_corr.dropna(how='all', axis=0).dropna(how='all', axis=1)
formatted_df = heatmap_data.fillna("").astype(str)

output_file = "data/formatted_correlation_matrix.xlsx"




with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    formatted_df.to_excel(writer,  sheet_name='Correlation Matrix')

    workbook = writer.book
    worksheet = workbook['Correlation Matrix']

    for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            if '-' in cell.value:
                cell.font = Font(bold=True)

    workbook.save(output_file)

print(f"Formatted correlation matrix saved to {output_file}")
